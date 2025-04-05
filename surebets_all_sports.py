import requests
import time
from datetime import datetime
from openpyxl import Workbook, load_workbook
import os

# === CONFIGURACIONES ===
API_KEY = '29351cd342f7b7a00601194d4e0fd3c2'
TELEGRAM_TOKEN = '7596593771:AAGAAmw_p1tnM09XlYQ5QN0G2Rjzoi61Sks'
TELEGRAM_CHAT_ID = '6382291298'
REGIONES = 'eu'
CASAS_VALIDAS = ['Bet365', 'Pinnacle', 'Betway', 'Unibet', 'Bwin']
EXCEL_FILE = 'surebets.xlsx'

# === ENVIAR MENSAJE A TELEGRAM ===
def enviar_telegram(mensaje):
    url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage"
    payload = {
        'chat_id': TELEGRAM_CHAT_ID,
        'text': mensaje,
        'parse_mode': 'HTML'
    }
    try:
        requests.post(url, data=payload)
    except Exception as e:
        print(f"❌ Error al enviar mensaje a Telegram: {e}")

# === GUARDAR EN EXCEL ===
def guardar_en_excel(fecha, deporte, partido, detalles, ganancia_pct):
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["Fecha", "Deporte", "Partido", "Detalles", "Ganancia (%)"])
        wb.save(EXCEL_FILE)

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([fecha, deporte, partido, detalles, ganancia_pct])
    wb.save(EXCEL_FILE)

# === OBTENER DEPORTES ACTIVOS ===
def obtener_deportes():
    url = 'https://api.the-odds-api.com/v4/sports/'
    res = requests.get(url, params={'apiKey': API_KEY})
    if res.status_code != 200:
        print(f"❌ Error al obtener deportes: {res.status_code}")
        return []
    return [deporte for deporte in res.json() if deporte['active']]

# === BUSCAR SUREBETS ===
contador_diario = []

def buscar_surebets(deporte_key, deporte_nombre):
    global contador_diario
    url = f'https://api.the-odds-api.com/v4/sports/{deporte_key}/odds/'
    params = {
        'apiKey': API_KEY,
        'regions': REGIONES,
        'markets': 'h2h',
        'oddsFormat': 'decimal'
    }
    res = requests.get(url, params=params)

    if res.status_code != 200:
        print(f"⚠️ No se pudo acceder a {deporte_nombre}")
        return

    eventos = res.json()
    for evento in eventos:
        home = evento.get('home_team')
        away = evento.get('away_team')
        nombre_partido = f"{home} vs {away}"
        mejores_momios = {}

        for casa in evento.get('bookmakers', []):
            nombre_casa = casa['title']
            if nombre_casa not in CASAS_VALIDAS:
                continue

            for cuota in casa['markets'][0]['outcomes']:
                participante = cuota['name']
                momio = cuota['price']
                if participante not in mejores_momios or momio > mejores_momios[participante]['momio']:
                    mejores_momios[participante] = {
                        'momio': momio,
                        'casa': nombre_casa
                    }

        if len(mejores_momios) < 2:
            continue

        suma_prob = sum(1 / datos['momio'] for datos in mejores_momios.values())
        if suma_prob < 1:
            ganancia_pct = round((1 - suma_prob) * 100, 2)
            print(f"\n✅ ¡HAY SUREBET! {nombre_partido}")
            mensaje = f"<b>🔥 ¡SUREBET!</b>\n{nombre_partido}\n"
            inversion_total = 500
            detalles = ""
            for equipo, datos in mejores_momios.items():
                inversion = round((inversion_total / datos['momio']) / suma_prob, 2)
                ganancia = round(inversion * datos['momio'], 2)
                print(f"💰 Apostar ${inversion} a {equipo} en {datos['casa']} (Retorno: ${ganancia})")
                mensaje += f"\n{equipo}: {datos['momio']} en {datos['casa']} - Apostar ${inversion}"
                detalles += f"{equipo}: {datos['momio']} en {datos['casa']} | "

            ganancia_neta = round((inversion_total / suma_prob) - inversion_total, 2)
            print(f"📈 Ganancia asegurada: ${ganancia_neta}")
            mensaje += f"\n\n📈 Ganancia estimada: <b>${ganancia_neta}</b> ({ganancia_pct}%)"
            enviar_telegram(mensaje)
            fecha_actual = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            guardar_en_excel(fecha_actual, deporte_nombre, nombre_partido, detalles, ganancia_pct)
            contador_diario.append(ganancia_pct)

# === ENVIAR RESUMEN DIARIO ===
def enviar_resumen_diario():
    if contador_diario:
        total = len(contador_diario)
        promedio = round(sum(contador_diario) / total, 2)
        mensaje = f"📊 Resumen del día:\nTotal de surebets: {total}\nGanancia promedio: {promedio}%"
        enviar_telegram(mensaje)
    else:
        enviar_telegram("📊 Hoy no se detectaron surebets.")

# === LOOP PRINCIPAL ===
if __name__ == "__main__":
    ultimo_dia = datetime.now().day
    while True:
        ahora = datetime.now()
        if ahora.day != ultimo_dia:
            enviar_resumen_diario()
            contador_diario = []
            ultimo_dia = ahora.day

        deportes = obtener_deportes()
        print(f"\n🔁 Revisión cada 3 minutos ({len(deportes)} deportes activos)...")
        for deporte in deportes:
            buscar_surebets(deporte['key'], deporte['title'])
            time.sleep(1)
        print("\n⏳ Esperando 3 minutos para próxima revisión...\n")
        time.sleep(180)


